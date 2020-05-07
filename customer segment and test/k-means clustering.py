# -*- coding:utf-8 -*-
import numpy as np
from matplotlib import pyplot
# excel写出和读取，后面可调用
import openpyxl
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")
def read_excel_xlsx(path, sheet_name):
    mylist=list()
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        item=list()
        for cell in row:
            #print(cell.value, "\t", end="")
            item.append(cell.value)
        mylist.append(item)
    return mylist

class K_Means(object):
    # k是分组数；tolerance‘中心点误差’；max_iter是迭代次数
    def __init__(self, k=2, tolerance=0.0001, max_iter=300):
        self.k_ = k
        self.tolerance_ = tolerance
        self.max_iter_ = max_iter

    def fit(self, data):
        self.centers_ = {}
        for i in range(self.k_):
            self.centers_[i] = data[i]

        for i in range(self.max_iter_):
            self.clf_ = {}
            for i in range(self.k_):
                self.clf_[i] = []
            # print("质点:",self.centers_)
            for feature in data:
                # distances = [np.linalg.norm(feature-self.centers[center]) for center in self.centers]
                distances = []
                for center in self.centers_:
                    # 欧拉距离
                    # np.sqrt(np.sum((features-self.centers_[center])**2))
                    distances.append(np.linalg.norm(feature - self.centers_[center]))
                classification = distances.index(min(distances))
                self.clf_[classification].append(feature)

            # print("分组情况:",self.clf_)
            prev_centers = dict(self.centers_)
            for c in self.clf_:
                self.centers_[c] = np.average(self.clf_[c], axis=0)

            # '中心点'是否在误差范围
            optimized = True
            for center in self.centers_:
                org_centers = prev_centers[center]
                cur_centers = self.centers_[center]
                if np.sum((cur_centers - org_centers) / org_centers * 100.0) > self.tolerance_:
                    optimized = False
            if optimized:
                break

    def predict(self, p_data):
        distances = [np.linalg.norm(p_data - self.centers_[center]) for center in self.centers_]
        index = distances.index(min(distances))
        return index


if __name__ == '__main__':

    # 读取excel
    book_name_xlsx = 'C:/Users/黎月明/Desktop/毕业论文/匹配.csv'
    sheet_name_xlsx = 's1'
    mylist = read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
    x = np.array(mylist)
    k_means = K_Means(k=2)
    k_means.fit(x)
    print(k_means.centers_)
    for center in k_means.centers_:
        pyplot.scatter(k_means.centers_[center][0], k_means.centers_[center][1], marker='*', s=150)

    for cat in k_means.clf_:
        for point in k_means.clf_[cat]:
            pyplot.scatter(point[0], point[1], c=('r' if cat == 0 else 'b'))

    pyplot.show()